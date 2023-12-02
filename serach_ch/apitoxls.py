import requests
from bs4 import BeautifulSoup
import json
from openpyxl.styles import PatternFill, Font, Color
import openpyxl

class ParserLocalCh:
    def __init__(self):
        self.session = requests.Session()
        self.data = {}
        try:
            self.parse()
            self.save_file_to_json()
        except:
             self.save_file_to_json()
             self.parse()
    def parse(self):
        
        response = requests.get("https://search.ch/tel/?kanton=VD&tel=07&privat=1&pages=20")

    # Check if the request was successful (status code 200)
    # Parse the HTML content with BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the relevant HTML elements using BeautifulSoup methods
        result_entries = soup.find_all('table', class_='tel-resultentry')

    # Iterate through the result entries
        for i, result_entry in enumerate(result_entries):
        # Extract information from each result entry
                name = result_entry.find('h1').text.strip()
                occupation_div = result_entry.find('div', class_='tel-occupation')
                occupation = occupation_div.text.strip() if occupation_div else None
                address_div = result_entry.find('div', class_='tel-address')
                address = address_div.text.strip()
                postal_code_span = address_div.find('span', class_='postal-code')
                postal_code = postal_code_span.text.strip() if postal_code_span else None
                phone_number = result_entry.find('a', class_='value').text.strip()
                entry_data = {
                "phone_number": [phone_number],  # You may need to modify this based on the actual structure
                "occupation": occupation,
                "adresse": address,
                "postal_code": postal_code,
                 }

                # Add the entry data to the self.data dictionary using the name as the key
                self.data[name] = entry_data
    # Print or process the extracted information as needed
                print('done'+str(i))
        print(self.data)

   
    def save_file_to_json(self):
        with open("data.json", "w", encoding="utf-8") as f:
            json.dump(self.data, f)

#parser=ParserLocalCh()

  

def save_json_to_excel(json_path, path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the fill color (grey)
    fill = PatternFill(fill_type="solid", fgColor=Color(rgb="00D3D3D3"))

    # Add headers
    headers = ['Nom Particulier', 'Ocupation', 'Phone Numbers', 'Adresse','code_Poastal']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 50

    # Add data
    for row_num, (company_name, company_data) in enumerate(data.items(), start=2):
        ws.cell(row=row_num, column=1, value=company_name)
        ws.cell(row=row_num, column=2, value=company_data.get('occupation'))
        ws.cell(row=row_num, column=3, value=", ".join(company_data.get('phone_number', [])))
        ws.cell(row=row_num, column=4, value=company_data.get('adresse'))
        ws.cell(row=row_num, column=5, value=company_data.get('postal_code'))

        # Alternate row color
        if row_num % 2 == 0:
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).fill = fill

    wb.save(path)


print('convering json to excel')
save_json_to_excel("data.json", "local-lausanne.-searchch.xlsx")