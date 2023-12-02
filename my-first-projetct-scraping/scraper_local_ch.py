import requests
from bs4 import BeautifulSoup
import json
from openpyxl.styles import PatternFill, Font, Color
import openpyxl

class ParserLocalCh:
    """
    This class will take from www.local.ch these informations:
        -Phone numbers that are in range (05-09) (two first digits)
        -Email
        -Adresse
    """
    def __init__(self):
        self.base_url = "https://www.local.ch/en/s/Lausanne%20(Region)?rid=56238b"
        self.session = requests.Session()
        self.data = {}
        self.last_page_index = self.get_last_page_index()
        try:
            self.parse()
            self.save_file_to_json()
        except:
            self.save_file_to_json()

    def save_file_to_json(self):
        with open("data.json", "w", encoding="utf-8") as f:
            json.dump(self.data, f)

    def get_last_page_index(self):
        request = self.session.get(self.base_url)
        soup = self.get_soup(request.text)
        ol_liste = soup.find("ol", {"class": "Pagination_pagesList__H30Dj"})
        lis = ol_liste.find_all("li")
        last_page_element = lis[-1]
        last_page_a = last_page_element.find("a").text
        return int(last_page_a)

    def parse(self, index_page=None):
        if index_page is None:
            print("Parsing first page")
            link_to_use = self.base_url
        else:
            link_to_use = self.base_url + f"&page={index_page}"
            print(f"Parsing page number {index_page}")

        links = self.get_links_from_page(link_to_use)
        for i, link_company in enumerate(links):
            print(f"\t Parsing link number {i+1}/{len(links)}")
            self.parse_company(link_company)
        if index_page is None:
            index_page = 2
        else:
            if index_page == self.last_page_index:
                return
            index_page = index_page+1
        self.parse(index_page)

    def get_links_from_page(self, url):
        request_to_website = self.session.get(url)
        soup = self.get_soup(request_to_website.text)
        all_links_companies_element = soup.find_all("a", {"class": "ListElement_link__LabW8"})
        liste_links = []
        for element in all_links_companies_element:
            link = "https://www.local.ch/" + element["href"]
            liste_links.append(link)
        return liste_links

    def parse_company(self, link_company):
        request_company = self.session.get(link_company)
        soup_company = self.get_soup(request_company.text)
        company_name = soup_company.find("h1", {"data-cy": "header-title"}).text
        self.data[company_name] = {}
        liste_numbers = []
        email = None
        contact_elements = soup_company.find_all("div", {"class": "ContactGroupsAccordion_contactContainer__yFs6g"})
        for contact_element in contact_elements:
            contact = contact_element.find("a").text
            contact = contact.replace("*", "")
            contact = contact.replace(" ", "")
            if self.is_intable(contact):
                if self.number_requirements(contact) is True:
                    liste_numbers.append(contact)
            else:
                if "@" in contact:
                    email = contact
        adresse = soup_company.find("a", {"class": "l--link DetailMapPreview_addressValue__pQROv"})
        if adresse is None:
            adresse = soup_company.find("span", {"class": "DetailMapPreview_addressValue__pQROv"})
        adresse = adresse.text
        self.data[company_name]["phone_numbers"] = liste_numbers
        self.data[company_name]["email"] = email
        self.data[company_name]["adresse"] = adresse

    def get_soup(self, html_text):
        return BeautifulSoup(html_text, 'html.parser')

    def is_intable(self, string):
        string = string.replace(" ", "")
        try:
            int(string)
            return True
        except:
            return False

    def number_requirements(self, number):
        two_first_digits = number[:2]
        if two_first_digits in ["05", "06", "07", "08", "09"]:
            return True
        return False


parser = ParserLocalCh()


def save_json_to_excel(json_path, path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the fill color (grey)
    fill = PatternFill(fill_type="solid", fgColor=Color(rgb="00D3D3D3"))

    # Add headers
    headers = ['Company Name', 'Email', 'Phone Numbers', 'Adresse']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 50

    # Add data
    for row_num, (company_name, company_data) in enumerate(data.items(), start=2):
        ws.cell(row=row_num, column=1, value=company_name)
        ws.cell(row=row_num, column=2, value=company_data.get('email'))
        ws.cell(row=row_num, column=3, value=", ".join(company_data.get('phone_numbers', [])))
        ws.cell(row=row_num, column=4, value=company_data.get('adresse'))

        # Alternate row color
        if row_num % 2 == 0:
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).fill = fill

    wb.save(path)


save_json_to_excel("data.json", "local-lausanne.xlsx")